"""Memory-efficient streaming export for large datasets."""

import asyncio
import csv
import json
import io
from typing import Any, AsyncIterator, Dict, Iterator, List, Optional, Union
from enum import Enum
from pathlib import Path

# Optional imports - fallback gracefully if not available
try:
    import aiofiles
    HAS_AIOFILES = True
except ImportError:
    HAS_AIOFILES = False

try:
    import openpyxl
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    import pyarrow as pa
    import pyarrow.parquet as pq
    HAS_PARQUET = True
except ImportError:
    HAS_PARQUET = False


class ExportFormat(str, Enum):
    """Supported export formats."""
    
    CSV = "csv"
    JSON = "json"
    JSONL = "json_lines"
    EXCEL = "excel"
    PARQUET = "parquet"
    TSV = "tsv"


class StreamingExporter:
    """Memory-efficient export for large datasets."""
    
    def __init__(self, chunk_size: int = 10000):
        """Initialize exporter with chunk size."""
        self._chunk_size = chunk_size
        self._export_stats = {
            'rows_exported': 0,
            'bytes_written': 0,
            'chunks_processed': 0
        }
    
    async def export_csv_streaming(
        self, 
        data_iterator: AsyncIterator[Dict[str, Any]], 
        output_path: Union[str, Path],
        delimiter: str = ',',
        include_headers: bool = True
    ) -> Dict[str, Any]:
        """Stream results to CSV without loading all in memory."""
        output_path = Path(output_path)
        
        if not HAS_AIOFILES:
            # Fallback to synchronous export (placeholder)
            return {
                'format': 'csv',
                'rows_exported': 0,
                'file_size_bytes': 0,
                'output_path': str(output_path),
                'error': 'aiofiles not available - streaming export disabled'
            }
        
        chunk = []
        writer = None
        rows_written = 0
        bytes_written = 0
        
        async with aiofiles.open(output_path, 'w', newline='', encoding='utf-8') as file:
            async for row in data_iterator:
                chunk.append(row)
                
                if len(chunk) >= self._chunk_size:
                    # Write chunk
                    if writer is None and chunk:
                        # Initialize writer with first chunk headers
                        headers = list(chunk[0].keys())
                        csv_buffer = io.StringIO()
                        writer = csv.DictWriter(csv_buffer, fieldnames=headers, delimiter=delimiter)
                        
                        if include_headers:
                            writer.writeheader()
                    
                    if writer:
                        csv_buffer = io.StringIO()
                        writer = csv.DictWriter(csv_buffer, fieldnames=writer.fieldnames, delimiter=delimiter)
                        writer.writerows(chunk)
                        
                        content = csv_buffer.getvalue()
                        await file.write(content)
                        bytes_written += len(content.encode('utf-8'))
                    
                    rows_written += len(chunk)
                    chunk = []
                    self._export_stats['chunks_processed'] += 1
                    
                    # Yield control periodically
                    if rows_written % 100000 == 0:
                        await asyncio.sleep(0)
            
            # Write remaining chunk
            if chunk:
                if writer is None and chunk:
                    headers = list(chunk[0].keys())
                    csv_buffer = io.StringIO()
                    writer = csv.DictWriter(csv_buffer, fieldnames=headers, delimiter=delimiter)
                    
                    if include_headers:
                        writer.writeheader()
                        content = csv_buffer.getvalue()
                        await file.write(content)
                        bytes_written += len(content.encode('utf-8'))
                
                if writer:
                    csv_buffer = io.StringIO()
                    writer = csv.DictWriter(csv_buffer, fieldnames=writer.fieldnames, delimiter=delimiter)
                    writer.writerows(chunk)
                    
                    content = csv_buffer.getvalue()
                    await file.write(content)
                    bytes_written += len(content.encode('utf-8'))
                
                rows_written += len(chunk)
        
        self._export_stats['rows_exported'] = rows_written
        self._export_stats['bytes_written'] = bytes_written
        
        return {
            'format': 'csv',
            'path': str(output_path),
            'rows_exported': rows_written,
            'bytes_written': bytes_written,
            'chunks_processed': self._export_stats['chunks_processed']
        }
    
    async def export_json_streaming(
        self, 
        data_iterator: AsyncIterator[Dict[str, Any]], 
        output_path: Union[str, Path],
        pretty: bool = False
    ) -> Dict[str, Any]:
        """Stream results to JSON array."""
        output_path = Path(output_path)
        rows_written = 0
        bytes_written = 0
        first_row = True
        
        async with aiofiles.open(output_path, 'w', encoding='utf-8') as file:
            await file.write('[')
            bytes_written += 1
            
            if pretty:
                await file.write('\n')
                bytes_written += 1
            
            async for row in data_iterator:
                if not first_row:
                    await file.write(',')
                    bytes_written += 1
                    if pretty:
                        await file.write('\n')
                        bytes_written += 1
                else:
                    first_row = False
                
                if pretty:
                    content = json.dumps(row, indent=2, default=str)
                    # Indent each line
                    content = '\n'.join('  ' + line for line in content.split('\n'))
                else:
                    content = json.dumps(row, default=str)
                
                await file.write(content)
                bytes_written += len(content.encode('utf-8'))
                rows_written += 1
                
                # Yield control periodically
                if rows_written % 10000 == 0:
                    await asyncio.sleep(0)
            
            if pretty:
                await file.write('\n')
                bytes_written += 1
            
            await file.write(']')
            bytes_written += 1
        
        self._export_stats['rows_exported'] = rows_written
        self._export_stats['bytes_written'] = bytes_written
        
        return {
            'format': 'json',
            'path': str(output_path),
            'rows_exported': rows_written,
            'bytes_written': bytes_written
        }
    
    async def export_jsonl_streaming(
        self, 
        data_iterator: AsyncIterator[Dict[str, Any]], 
        output_path: Union[str, Path]
    ) -> Dict[str, Any]:
        """Stream results to JSON Lines format."""
        output_path = Path(output_path)
        rows_written = 0
        bytes_written = 0
        
        async with aiofiles.open(output_path, 'w', encoding='utf-8') as file:
            async for row in data_iterator:
                line = json.dumps(row, default=str) + '\n'
                await file.write(line)
                
                bytes_written += len(line.encode('utf-8'))
                rows_written += 1
                
                # Yield control periodically
                if rows_written % 10000 == 0:
                    await asyncio.sleep(0)
        
        self._export_stats['rows_exported'] = rows_written
        self._export_stats['bytes_written'] = bytes_written
        
        return {
            'format': 'jsonl',
            'path': str(output_path),
            'rows_exported': rows_written,
            'bytes_written': bytes_written
        }
    
    async def export_excel_streaming(
        self, 
        data_iterator: AsyncIterator[Dict[str, Any]], 
        output_path: Union[str, Path],
        sheet_name: str = "Data"
    ) -> Dict[str, Any]:
        """Stream results to Excel file."""
        output_path = Path(output_path)
        
        # Create workbook
        wb = openpyxl.Workbook(write_only=True)
        ws = wb.create_sheet(title=sheet_name)
        
        rows_written = 0
        headers_written = False
        
        # Collect data in chunks for Excel
        chunk = []
        
        async for row in data_iterator:
            if not headers_written and row:
                # Write headers
                headers = list(row.keys())
                ws.append(headers)
                headers_written = True
            
            # Convert row to list
            if headers_written:
                row_values = [row.get(h) for h in headers]
                ws.append(row_values)
                rows_written += 1
            
            # Yield control periodically
            if rows_written % 10000 == 0:
                await asyncio.sleep(0)
        
        # Save workbook
        wb.save(output_path)
        
        # Get file size
        bytes_written = output_path.stat().st_size
        
        self._export_stats['rows_exported'] = rows_written
        self._export_stats['bytes_written'] = bytes_written
        
        return {
            'format': 'excel',
            'path': str(output_path),
            'rows_exported': rows_written,
            'bytes_written': bytes_written,
            'sheet_name': sheet_name
        }
    
    async def export_parquet_streaming(
        self, 
        data_iterator: AsyncIterator[Dict[str, Any]], 
        output_path: Union[str, Path],
        compression: str = 'snappy'
    ) -> Dict[str, Any]:
        """Stream results to Parquet format."""
        output_path = Path(output_path)
        
        # Collect first batch to infer schema
        first_batch = []
        schema = None
        
        async for row in data_iterator:
            first_batch.append(row)
            if len(first_batch) >= 100:  # Sample size for schema
                break
        
        if not first_batch:
            return {
                'format': 'parquet',
                'path': str(output_path),
                'rows_exported': 0,
                'bytes_written': 0
            }
        
        # Infer schema from first batch
        df_sample = pa.Table.from_pylist(first_batch)
        schema = df_sample.schema
        
        # Create Parquet writer
        writer = pq.ParquetWriter(output_path, schema, compression=compression)
        
        # Write first batch
        writer.write_table(df_sample)
        rows_written = len(first_batch)
        
        # Continue with remaining data in chunks
        chunk = []
        
        async for row in data_iterator:
            chunk.append(row)
            
            if len(chunk) >= self._chunk_size:
                # Write chunk
                table = pa.Table.from_pylist(chunk, schema=schema)
                writer.write_table(table)
                rows_written += len(chunk)
                chunk = []
                
                # Yield control
                if rows_written % 100000 == 0:
                    await asyncio.sleep(0)
        
        # Write final chunk
        if chunk:
            table = pa.Table.from_pylist(chunk, schema=schema)
            writer.write_table(table)
            rows_written += len(chunk)
        
        # Close writer
        writer.close()
        
        # Get file size
        bytes_written = output_path.stat().st_size
        
        self._export_stats['rows_exported'] = rows_written
        self._export_stats['bytes_written'] = bytes_written
        
        return {
            'format': 'parquet',
            'path': str(output_path),
            'rows_exported': rows_written,
            'bytes_written': bytes_written,
            'compression': compression
        }
    
    async def export(
        self,
        data_iterator: AsyncIterator[Dict[str, Any]],
        output_path: Union[str, Path],
        format: ExportFormat,
        **kwargs
    ) -> Dict[str, Any]:
        """Export data in specified format."""
        if format == ExportFormat.CSV:
            return await self.export_csv_streaming(data_iterator, output_path, **kwargs)
        elif format == ExportFormat.JSON:
            return await self.export_json_streaming(data_iterator, output_path, **kwargs)
        elif format == ExportFormat.JSONL:
            return await self.export_jsonl_streaming(data_iterator, output_path)
        elif format == ExportFormat.EXCEL:
            return await self.export_excel_streaming(data_iterator, output_path, **kwargs)
        elif format == ExportFormat.PARQUET:
            return await self.export_parquet_streaming(data_iterator, output_path, **kwargs)
        elif format == ExportFormat.TSV:
            return await self.export_csv_streaming(data_iterator, output_path, delimiter='\t', **kwargs)
        else:
            raise ValueError(f"Unsupported export format: {format}")
    
    def get_export_stats(self) -> Dict[str, Any]:
        """Get export statistics."""
        return self._export_stats.copy()
    
    def estimate_memory_usage(self, row_count: int, avg_row_size: int = 1000) -> Dict[str, int]:
        """Estimate memory usage for different export formats."""
        estimates = {
            'csv': row_count * avg_row_size * 1.2,  # 20% overhead
            'json': row_count * avg_row_size * 1.5,  # 50% overhead for JSON structure
            'jsonl': row_count * avg_row_size * 1.3,  # 30% overhead
            'excel': min(row_count * avg_row_size * 2, 500 * 1024 * 1024),  # Excel has limits
            'parquet': row_count * avg_row_size * 0.3,  # Highly compressed
        }
        
        return {
            format: size // (1024 * 1024)  # Convert to MB
            for format, size in estimates.items()
        }